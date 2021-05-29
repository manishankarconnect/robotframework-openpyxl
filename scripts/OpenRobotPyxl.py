import openpyxl
from robot.api.deco import keyword, library


@library
class OpenRobotPyxl:

    def __init__(self):
        self.active_sheet = None
        self.active_book = None
        self.path = None
        self.bookname = None
        pass

    @keyword("Create New Workbook")
    def create_new_workbook(self, _path, book_name, sheet_name, postion=0):
        wb = openpyxl.Workbook()
        self.path = _path
        self.bookname = book_name + ".xlsx"
        ws = wb.create_sheet(sheet_name, postion)
        self.active_book, self.active_sheet = wb, ws
        return self.active_book

    @keyword('Close Workbook')
    def close_workbook(self):
        self.active_book.save(self.path + "/" + self.bookname)

    @keyword('Get Active Sheet')
    def get_active_sheet(self):
        if self.active_book:
            if self.active_sheet:
                return self.active_sheet
            else:
                # Return the first sheet in the work book.
                return self.active_book.worksheets[0]
        else:
            return None

    @keyword('Active Sheet Name')
    def get_active_sheet_name(self):
        return self.get_active_sheet().title

    @keyword('Load Workbook')
    def load_workbook(self, bookname):
        self.active_book = openpyxl.load_workbook(bookname)
        self.bookname = bookname
        self.active_sheet = self.get_active_sheet()



